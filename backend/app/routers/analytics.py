"""Analytics API:招聘漏斗与运营指标聚合 + Excel 导出(admin)。"""

import datetime
import io

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import Date, cast, distinct, func, select
from sqlalchemy.orm import Session

from ..auth import require_admin
from ..database import get_db
from ..models import (
    Application,
    Candidate,
    EmailLog,
    Interview,
    Interviewer,
    Job,
    Score,
    Slot,
    SlotInterviewer,
    UserSession,
)


def _analytics_payload(db: Session) -> dict:
    """聚合逻辑供 JSON 端点与 Excel 导出共用。"""
    return _compute_analytics(db)

router = APIRouter(prefix="/analytics", tags=["analytics"])


@router.get("")
def analytics(db: Session = Depends(get_db),
              _admin: UserSession = Depends(require_admin)) -> dict:
    return _compute_analytics(db)


def _compute_analytics(db: Session) -> dict:
    today = datetime.date.today()

    total_apps = db.scalar(select(func.count()).select_from(Application)) or 0
    total_candidates = db.scalar(select(func.count()).select_from(Candidate)) or 0
    open_jobs = db.scalar(select(func.count()).select_from(Job).where(Job.is_open)) or 0

    status_counts = dict(
        db.execute(select(Application.status, func.count()).group_by(Application.status)).all()
    )
    band_counts = dict(
        db.execute(select(Score.band, func.count()).group_by(Score.band)).all()
    )
    rejection_reasons = dict(
        db.execute(
            select(Application.rejected_reason, func.count())
            .where(Application.status == "rejected", Application.rejected_reason.isnot(None))
            .group_by(Application.rejected_reason)
        ).all()
    )

    invite_sent = db.scalar(
        select(func.count(distinct(EmailLog.application_id))).where(
            EmailLog.type == "invite", EmailLog.status == "sent"
        )
    ) or 0
    booked = db.scalar(select(func.count(distinct(Interview.application_id)))) or 0
    outcome_done = db.scalar(
        select(func.count()).select_from(Interview).where(Interview.status.in_(["passed", "failed"]))
    ) or 0
    passed = status_counts.get("passed", 0)

    # 漏斗:各阶段"到达过"的申请数(累进)
    screened_ok = (band_counts.get("high", 0) or 0) + (band_counts.get("medium", 0) or 0)
    funnel = [
        {"stage": "Applications received", "count": total_apps},
        {"stage": "Passed screening (High/Medium)", "count": screened_ok},
        {"stage": "Invite sent", "count": invite_sent},
        {"stage": "Interview booked", "count": booked},
        {"stage": "Outcome recorded", "count": outcome_done},
        {"stage": "Passed (offer)", "count": passed},
    ]

    per_job = [
        {"title": title, "applications": apps, "passed": passed_n}
        for title, apps, passed_n in db.execute(
            select(
                Job.title,
                func.count(Application.id),
                func.count(Application.id).filter(Application.status == "passed"),
            )
            .join(Application, Application.job_id == Job.id, isouter=True)
            .group_by(Job.id, Job.title)
            .order_by(func.count(Application.id).desc())
        ).all()
    ]

    since = today - datetime.timedelta(days=13)
    daily_rows = dict(
        db.execute(
            select(cast(Application.submitted_at, Date), func.count())
            .where(cast(Application.submitted_at, Date) >= since)
            .group_by(cast(Application.submitted_at, Date))
        ).all()
    )
    daily = [
        {"date": (since + datetime.timedelta(days=i)).isoformat(),
         "count": daily_rows.get(since + datetime.timedelta(days=i), 0)}
        for i in range(14)
    ]

    slot_counts = dict(
        db.execute(
            select(Slot.status, func.count())
            .where(Slot.slot_date >= today)
            .group_by(Slot.status)
        ).all()
    )

    interviewer_load = [
        {"name": name, "claimed": claimed, "booked": booked_n}
        for name, claimed, booked_n in db.execute(
            select(
                Interviewer.name,
                func.count(SlotInterviewer.slot_id),
                func.count(SlotInterviewer.slot_id).filter(Slot.status == "booked"),
            )
            .join(SlotInterviewer, SlotInterviewer.interviewer_id == Interviewer.id, isouter=True)
            .join(Slot, Slot.id == SlotInterviewer.slot_id, isouter=True)
            .group_by(Interviewer.id, Interviewer.name)
            .order_by(func.count(SlotInterviewer.slot_id).desc())
        ).all()
    ]

    # 提交到预约的平均天数
    avg_secs = db.scalar(
        select(func.avg(func.extract("epoch", Interview.created_at - Application.submitted_at)))
        .select_from(Interview)
        .join(Application, Application.id == Interview.application_id)
    )
    avg_days_to_book = round(float(avg_secs) / 86400, 1) if avg_secs is not None else None

    return {
        "overview": {
            "total_applications": total_apps,
            "total_candidates": total_candidates,
            "open_jobs": open_jobs,
            "offers": passed,
            "avg_days_to_book": avg_days_to_book,
        },
        "funnel": funnel,
        "bands": {b: band_counts.get(b, 0) for b in ("high", "medium", "low")},
        "statuses": status_counts,
        "rejection_reasons": rejection_reasons,
        "per_job": per_job,
        "daily_applications": daily,
        "slots": {
            "open": slot_counts.get("open", 0),
            "booked": slot_counts.get("booked", 0),
            "empty": slot_counts.get("empty", 0),
        },
        "interviewer_load": interviewer_load,
    }


# ---------- Excel 导出 ----------

@router.get("/export")
def export_excel(db: Session = Depends(get_db),
                 _admin: UserSession = Depends(require_admin)) -> StreamingResponse:
    """汇总指标 + 原始明细导出为多 sheet 的 .xlsx,供外部分析。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    a = _compute_analytics(db)
    wb = Workbook()
    bold = Font(bold=True)

    def sheet(title: str, headers: list[str], rows: list[list]) -> None:
        ws = wb.create_sheet(title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
        for row in rows:
            ws.append(row)
        # 简单列宽自适应
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(width + 2, 50)

    # Overview
    ws = wb.active
    ws.title = "Overview"
    ov = a["overview"]
    for k, v in [
        ("Generated at", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total applications", ov["total_applications"]),
        ("Total candidates", ov["total_candidates"]),
        ("Open jobs", ov["open_jobs"]),
        ("Offers (passed)", ov["offers"]),
        ("Avg days to book", ov["avg_days_to_book"] if ov["avg_days_to_book"] is not None else "n/a"),
    ]:
        ws.append([k, v])
    for row in ws.iter_rows(min_col=1, max_col=1):
        row[0].font = bold
    ws.column_dimensions["A"].width = 24

    sheet("Funnel", ["Stage", "Count"], [[f["stage"], f["count"]] for f in a["funnel"]])
    sheet("Breakdowns", ["Category", "Key", "Count"],
          [["band", b, n] for b, n in a["bands"].items()]
          + [["status", k, v] for k, v in a["statuses"].items()]
          + [["rejection_reason", k, v] for k, v in a["rejection_reasons"].items()])
    sheet("Daily applications", ["Date", "Count"],
          [[d["date"], d["count"]] for d in a["daily_applications"]])

    # Applications 原始明细
    app_rows = db.execute(
        select(Application, Candidate, Job, Score)
        .join(Candidate, Candidate.id == Application.candidate_id)
        .join(Job, Job.id == Application.job_id)
        .outerjoin(Score, Score.application_id == Application.id)
        .order_by(Application.submitted_at)
    ).all()
    rows = []
    for ap, c, j, sc in app_rows:
        fd = ap.form_data or {}
        jd = (ap.resume_parsed or {}).get("jd_match") or {}
        rows.append([
            ap.submitted_at.strftime("%Y-%m-%d %H:%M"),
            c.name, c.email, j.title,
            fd.get("education_level") or "",
            ap.degree_field or "", fd.get("institution") or "",
            float(ap.cgpa) if ap.cgpa is not None else None,
            "yes" if ap.is_fulltime else "no",
            ", ".join(ap.prog_langs or []),
            "yes" if ap.has_sql else "no",
            "yes" if ap.has_ai_study else "no",
            ", ".join(fd.get("skills") or []),
            sc.band if sc else "",
            float(sc.total_score) if sc and sc.total_score is not None else None,
            jd.get("match_score"),
            ap.status, ap.rejected_reason or "",
            "yes" if ap.resume_file_url else "no",
            fd.get("preferred_start_date") or "",
            fd.get("salary_expectation") or "",
            fd.get("heard_about_us") or "",
        ])
    sheet("Applications", [
        "Submitted", "Candidate", "Email", "Job", "Education level", "Field",
        "Institution", "CGPA", "Full-time", "Languages", "SQL", "AI",
        "Skills", "Band", "Score", "JD match", "Status", "Rejected reason",
        "Resume", "Preferred start", "Salary expectation", "Source",
    ], rows)

    # Interviews 明细
    itv_rows = db.execute(
        select(Interview, Slot, Application, Candidate, Job)
        .join(Slot, Slot.id == Interview.slot_id)
        .join(Application, Application.id == Interview.application_id)
        .join(Candidate, Candidate.id == Application.candidate_id)
        .join(Job, Job.id == Application.job_id)
        .order_by(Slot.slot_date, Slot.start_time)
    ).all()
    panel_map: dict = {}
    for sid, name in db.execute(
        select(SlotInterviewer.slot_id, Interviewer.name)
        .join(Interviewer, Interviewer.id == SlotInterviewer.interviewer_id)
    ).all():
        panel_map.setdefault(sid, []).append(name)
    sheet("Interviews", ["Date", "Time", "Candidate", "Job", "Panel", "Status", "Reschedules"],
          [[sl.slot_date.isoformat(),
            f"{sl.start_time.strftime('%H:%M')}-{sl.end_time.strftime('%H:%M')}",
            c.name, j.title, ", ".join(panel_map.get(sl.id, [])),
            itv.status, itv.reschedule_count]
           for itv, sl, ap, c, j in itv_rows])

    sheet("Interviewer load", ["Interviewer", "Claimed slots", "Booked"],
          [[i["name"], i["claimed"], i["booked"]] for i in a["interviewer_load"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"has-analytics-{datetime.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
